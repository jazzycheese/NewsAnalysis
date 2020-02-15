#엑셀에서 url가져오기
from openpyxl import load_workbook
import requests
from openpyxl import Workbook

load_wb = load_workbook("k_NewsURL.xlsx", data_only=True)
load_ws = load_wb['Sheet']

write_wb = Workbook()
write_ws = write_wb.active

print('\n-----지정한 셀 출력-----')
get_cells = load_ws['A1':'B1']   #['A1':'FS1']
for row in get_cells:
    for cell in row:
        print(cell.value)       #print (type(cell.value)) <class 'str'>
        s = str(cell.value)
        response = requests.get(s)
        html = response.text
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, 'html.parser')
        for tag in soup.select('.on a'):
            print(tag.text) 
            write_ws.append([tag.text])
            write_wb.save('k_Result.xlsx')
print('hello world')
#for tag in soup.select("p", align_="justify"):
#        print(tag.text)
#엑셀에서 url가져오기
from openpyxl import load_workbook
import requests
from openpyxl import Workbook

load_wb = load_workbook("k_NewsURL.xlsx", data_only=True)
load_ws = load_wb['Sheet']

write_wb = Workbook()
write_ws = write_wb.active

print('\n-----지정한 셀 출력-----')
get_cells = load_ws['A1':'FS1']  #FS1
for row in get_cells:
    for cell in row:
        print(cell.value)       #print (type(cell.value)) <class 'str'>
        s = str(cell.value)

        agent = {"User-Agent":'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36'}
        response = requests.get(s, headers=agent)

        html = response.text
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, 'html.parser')
        
        for tag in soup.select('p', class_='content_text'):
            print(tag.text) 
            write_ws.append([tag.text])
            write_wb.save('k_Result.xlsx')


#for tag in soup.select("p", align_="justify"):
#        print(tag.text)